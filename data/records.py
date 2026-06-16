"""Load legal case rows from the task spreadsheet."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

_CASE_TYPE_MAP = {
    "Regulatory Compliance": "Compliance",
}


@dataclass(frozen=True)
class CaseRecord:
    index: int
    case_ref: str
    client_name: str
    case_type: str
    jurisdiction: str
    record_source: str
    document_status: str
    case_status: str
    case_summary: str

    @property
    def case_type_value(self) -> str:
        return _CASE_TYPE_MAP.get(self.case_type, self.case_type)


def load_task_file(path: Path) -> list[CaseRecord]:
    """Read case rows from the downloaded task workbook."""
    if not path.is_file():
        raise FileNotFoundError(f"Task file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ValueError(f"Task file is empty: {path}")

    records: list[CaseRecord] = []
    for offset, row in enumerate(rows[1:], start=0):
        if not row or not row[0]:
            continue
        records.append(
            CaseRecord(
                index=offset,
                case_ref=str(row[0]).strip(),
                client_name=str(row[1] or "").strip(),
                case_type=str(row[2] or "").strip(),
                jurisdiction=str(row[3] or "").strip(),
                record_source=str(row[4] or "").strip(),
                document_status=str(row[5] or "Draft").strip(),
                case_status=str(row[6] or "Open").strip(),
                case_summary=str(row[7] or "").strip(),
            )
        )

    if not records:
        raise ValueError(f"No data rows found in task file: {path}")

    return records
